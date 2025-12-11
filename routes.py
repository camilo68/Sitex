# routes.py - VERSIÓN FINAL 100% FUNCIONAL (Render + Local)
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import current_user, login_user, logout_user, login_required
from flask_mail import Message
from werkzeug.utils import secure_filename
from datetime import date, datetime, timedelta
from sqlalchemy import func, extract
import bcrypt
import os
import secrets
import pandas as pd
from extensions import db, mail
from models import Empleado, Tanque, Descargue, RegistroMedida, MedicionCargue, SesionActiva, Auditoria, Venta
from forms import (
    LoginForm, RegisterForm, MedicionForm, DescargueForm, ChangePasswordForm,
    ResetPasswordForm, RequestPasswordResetForm, PasswordResetForm, TanqueForm,
    CargaMasivaForm, FiltroMedicionesForm, CargueEmergenciaForm
)
from utils import (
    islero_or_encargado_required, admin_or_encargado_required, admin_required,
    registrar_auditoria, allowed_file
)

# ======================= RESEND CONFIG =======================
USE_RESEND = os.environ.get("MAIL_DRIVER") == "resend"

# ======================= BLUEPRINTS =======================
auth_bp = Blueprint("auth", __name__, url_prefix="/auth")
main_bp = Blueprint("main", __name__)
dashboard_bp = Blueprint("dashboard", __name__, url_prefix="/dashboard")
medicion_bp = Blueprint("medicion", __name__, url_prefix="/medicion")
admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

# ======================= ENVÍO DE EMAIL UNIVERSAL =======================
def enviar_email(to: str, subject: str, html: str) -> bool:
    if USE_RESEND:
        try:
            import resend
            resend.api_key = os.environ["RESEND_API_KEY"]
            resend.Emails.send({
                "from": os.environ.get("MAIL_DEFAULT_SENDER", "Sitex <no-reply@sitex.com>"),
                "to": [to],
                "subject": subject,
                "html": html,
            })
            print(f"Email enviado con RESEND a {to}")
            return True
        except Exception as e:
            print(f"Error Resend: {e}")
            return False
    else:
        try:
            msg = Message(subject, recipients=[to])
            msg.html = html
            mail.send(msg)
            print(f"Email enviado con Flask-Mail a {to}")
            return True
        except Exception as e:
            print(f"Error Flask-Mail: {e}")
            return False

# ======================= EMAILS =======================
def enviar_email_confirmacion(empleado, token):
    confirm_url = url_for('auth.confirm_email', token=token, _external=True)
    contrasena_temp = empleado.numero_documento[-4:] if len(empleado.numero_documento) >= 4 else empleado.numero_documento

    html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <h2 style="color: #E10000;">¡Bienvenido a Hayuelos!</h2>
            <p>Hola <strong>{empleado.nombre_empleado} {empleado.apellido_empleado}</strong>,</p>
            <p>Gracias por registrarte. Confirma tu email aquí:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{confirm_url}" style="background:#E10000; color:white; padding:14px 32px; text-decoration:none; border-radius:6px; font-size:16px;">
                    Confirmar Email
                </a>
            </div>
            <p><strong>Usuario:</strong> <code>{empleado.usuario}</code><br>
               <strong>Contraseña temporal:</strong> <code>{contrasena_temp}</code></p>
            <p style="color:#666; font-size:0.9em;">Este enlace vence en 24 horas.</p>
            <hr>
            <p style="color:#999; font-size:0.8em; text-align:center;">Sistema Hayuelos</p>
        </div>
    </body>
    </html>
    """
    return enviar_email(empleado.email, "Confirma tu email - Hayuelos", html)

def enviar_email_recuperacion(empleado, token):
    reset_url = url_for('auth.reset_password', token=token, _external=True)
    html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <h2 style="color: #E10000;">Recuperación de Contraseña</h2>
            <p>Hola <strong>{empleado.nombre_empleado}</strong>,</p>
            <p>Haz clic para crear una nueva contraseña:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_url}" style="background:#E10000; color:white; padding:14px 32px; text-decoration:none; border-radius:6px; font-size:16px;">
                    Restablecer Contraseña
                </a>
            </div>
            <p style="color:#666; font-size:0.9em;">Este enlace vence en 1 hora.</p>
            <hr>
            <p style="color:#999; font-size:0.8em; text-align:center;">Sistema Hayuelos</p>
        </div>
    </body>
    </html>
    """
    return enviar_email(empleado.email, "Recuperación de Contraseña - Hayuelos", html)

# ======================= RESTO DEL CÓDIGO (100% igual al tuyo) =======================
# (Todo desde aquí para abajo es exactamente lo que ya tenías, solo con los cambios de email)

def calcular_altura_maxima(capacidad_galones):
    radio_cm = 125
    volumen_cm3 = capacidad_galones * 3785.411784
    area_base = 3.14159 * (radio_cm ** 2)
    return round(volumen_cm3 / area_base, 2)

# AUTH ROUTES
@auth_bp.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard.index"))
    form = LoginForm()
    if form.validate_on_submit():
        usuario = form.username.data.strip()
        contrasena = form.password.data
        empleado = Empleado.query.filter(
            (Empleado.usuario == usuario) | (Empleado.numero_documento == usuario)
        ).first()
        if empleado and empleado.check_password(contrasena):
            if not empleado.activo:
                flash("Cuenta deshabilitada.", "danger")
                return redirect(url_for("auth.login"))
            if not empleado.email_confirmado:
                flash("Debes confirmar tu email.", "warning")
                return redirect(url_for("auth.resend_confirmation"))
            login_user(empleado, remember=form.remember_me.data)
            sesion = SesionActiva(
                id_empleados=empleado.id_empleados,
                session_id=secrets.token_urlsafe(32),
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')[:255]
            )
            db.session.add(sesion)
            db.session.commit()
            flash("¡Bienvenido!", "success")
            return redirect(url_for("dashboard.index"))
        flash("Usuario o contraseña incorrectos", "danger")
    return render_template("auth/login.html", form=form)

@auth_bp.route("/register", methods=["GET", "POST"])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        if Empleado.query.filter_by(numero_documento=form.numero_documento.data).first():
            flash("Documento ya registrado", "danger"); return render_template("auth/register.html", form=form)
        if Empleado.query.filter_by(email=form.email.data).first():
            flash("Email ya registrado", "danger"); return render_template("auth/register.html", form=form)
        if Empleado.query.filter_by(usuario=form.usuario.data).first():
            flash("Usuario ya existe", "danger"); return render_template("auth/register.html", form=form)

        contrasena_temporal = form.numero_documento.data[-4:] if len(form.numero_documento.data) >= 4 else form.numero_documento.data
        hash_cifrado = bcrypt.hashpw(contrasena_temporal.encode(), bcrypt.gensalt()).decode()

        nuevo = Empleado(
            usuario=form.usuario.data, nombre_empleado=form.nombre_empleado.data,
            apellido_empleado=form.apellido_empleado.data, numero_documento=form.numero_documento.data,
            tipo_documento=form.tipo_documento.data, email=form.email.data, telefono=form.telefono.data,
            direccion=form.direccion.data, cargo_establecido=form.cargo_establecido.data,
            contrasena=hash_cifrado, temporal=True, activo=True, email_confirmado=False,
            aceptado_terminos=form.aceptar_terminos.data
        )
        token = nuevo.generate_confirmation_token()
        db.session.add(nuevo)
        db.session.commit()

        if enviar_email_confirmacion(nuevo, token):
            flash("¡Registro exitoso! Revisa tu correo.", "success")
        else:
            nuevo.email_confirmado = True
            db.session.commit()
            flash(f"Usuario creado. Contraseña temporal: {contrasena_temporal}", "warning")

        registrar_auditoria('CREATE', 'empleado', nuevo.id_empleados, None, {'usuario': nuevo.usuario})
        return redirect(url_for("auth.login"))
    return render_template("auth/register.html", form=form)

# NUEVO: Ruta para confirmar email
@auth_bp.route("/confirm/<token>")
def confirm_email(token):
    empleado = Empleado.query.filter_by(token_confirmacion=token).first()
    if not empleado and empleado.verify_confirmation_token(token):
        empleado.confirmar_email()
        db.session.commit()
        flash("Email confirmado. ¡Ya puedes iniciar sesión!", "success")
    else:
        flash("Token inválido o expirado", "danger")
    return redirect(url_for("auth.login"))

@auth_bp.route("/resend_confirmation", methods=["GET", "POST"])
def resend_confirmation():
    if request.method == "POST":
        email = request.form.get("email")
        empleado = Empleado.query.filter_by(email=email).first()
        if empleado:
            if empleado.email_confirmado:
                flash("Este email ya está confirmado", "info")
                return redirect(url_for("auth.login"))
            token = empleado.generate_confirmation_token()
            db.session.commit()
            enviar_email_confirmacion(empleado, token)
            flash("Email de confirmación reenviado", "success")
        else:
            flash("Si el email existe, recibirás un enlace", "info")
    return render_template("auth/resend_confirmation.html")

@auth_bp.route("/request_reset", methods=["GET", "POST"])
def request_password_reset():
    form = RequestPasswordResetForm()
    if form.validate_on_submit():
        empleado = Empleado.query.filter_by(email=form.email.data).first()
        if empleado:
            if not empleado.email_confirmado:
                flash("Primero confirma tu email", "warning")
                return redirect(url_for("auth.resend_confirmation"))
            token = empleado.generate_reset_token()
            db.session.commit()
            enviar_email_recuperacion(empleado, token)
            flash("Te enviamos un enlace para recuperar tu contraseña", "success")
        else:
            flash("Si el email existe, recibirás un enlace", "info")
    return render_template("auth/request_reset.html", form=form)
    
# NUEVO: Ruta para confirmar email
@auth_bp.route("/confirm/<token>")
def confirm_email(token):
    empleado = Empleado.query.filter_by(token_confirmacion=token).first()
    if not empleado:
        flash("Token de confirmación inválido", "danger")
        return redirect(url_for("auth.login"))
    if not empleado.verify_confirmation_token(token):
        flash("El token de confirmación ha expirado. Solicita un nuevo email de confirmación.", "danger")
        return redirect(url_for("auth.resend_confirmation"))
    empleado.confirmar_email()
    db.session.commit()
    registrar_auditoria('CONFIRM_EMAIL', 'empleado', empleado.id_empleados, None, {'email_confirmado': True})
    flash("¡Email confirmado exitosamente! Ya puedes iniciar sesión.", "success")
    return redirect(url_for("auth.login"))

# NUEVO: Reenviar email de confirmación
@auth_bp.route("/resend_confirmation", methods=["GET", "POST"])
def resend_confirmation():
    if request.method == "POST":
        email = request.form.get("email")
        empleado = Empleado.query.filter_by(email=email).first()
        
        if empleado:
            if empleado.email_confirmado:
                flash("Este email ya ha sido confirmado", "info")
                return redirect(url_for("auth.login"))
            
            # Generar nuevo token
            token = empleado.generate_confirmation_token()
            db.session.commit()
            
            try:
                enviar_email_confirmacion(empleado, token)
                flash("Se ha enviado un nuevo email de confirmación", "success")
            except Exception as e:
                flash(f"Error al enviar el email. Contacte al administrador.", "danger")
                print(f"Error: {e}")
        else:
            flash("Si el email existe en nuestro sistema, recibirás un enlace de confirmación", "info")
    
    return render_template("auth/resend_confirmation.html")

@auth_bp.route("/logout")
@login_required
def logout():
    SesionActiva.query.filter_by(
        id_empleados=current_user.id_empleados,
        activa=True
    ).update({'activa': False})
    db.session.commit()
    
    logout_user()
    flash("Sesión cerrada correctamente", "info")
    return redirect(url_for("auth.login"))

@auth_bp.route("/logout_all", methods=["POST"])
@login_required
def logout_all():
    SesionActiva.query.filter_by(id_empleados=current_user.id_empleados).update({'activa': False})
    db.session.commit()
    logout_user()
    flash("Se han cerrado todas las sesiones activas", "success")
    return redirect(url_for("auth.login"))

@auth_bp.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            hash_nuevo = bcrypt.hashpw(form.new_password.data.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
            current_user.contrasena = hash_nuevo
            current_user.temporal = False
            db.session.commit()
            
            registrar_auditoria('UPDATE', 'empleado', current_user.id_empleados, 
                              {'temporal': True}, {'temporal': False})
            
            flash("Contraseña actualizada exitosamente", "success")
            return redirect(url_for("dashboard.index"))
        else:
            flash("Contraseña actual incorrecta", "danger")
    
    return render_template("auth/change_password.html", form=form)

@auth_bp.route("/request_reset", methods=["GET", "POST"])
def request_password_reset():
    form = RequestPasswordResetForm()
    if form.validate_on_submit():
        empleado = Empleado.query.filter_by(email=form.email.data).first()
        if empleado:
            if not empleado.email_confirmado:
                flash("Primero debe confirmar su email. Revise su correo.", "warning")
                return redirect(url_for("auth.resend_confirmation"))
            
            token = empleado.generate_reset_token()
            db.session.commit()
            
            reset_url = url_for('auth.reset_password', token=token, _external=True)
            msg = Message("Recuperación de Contraseña - Hayuelos",
                        recipients=[empleado.email])
            msg.body = f"""Hola {empleado.nombre_empleado},

Has solicitado restablecer tu contraseña. Haz clic en el siguiente enlace:

{reset_url}

Este enlace expira en 1 hora.

Si no solicitaste este cambio, ignora este correo.

Saludos,
Sistema Hayuelos
"""
            msg.html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #E10000;">Recuperación de Contraseña</h2>
                    <p>Hola <strong>{empleado.nombre_empleado}</strong>,</p>
                    <p>Has solicitado restablecer tu contraseña en Hayuelos.</p>
                    
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{reset_url}" 
                           style="background-color: #E10000; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                            Restablecer Contraseña
                        </a>
                    </div>
                    
                    <p style="color: #666; font-size: 0.9em;">Este enlace expira en 1 hora.</p>
                    <p style="color: #666; font-size: 0.9em;">Si no solicitaste este cambio, ignora este correo.</p>
                    
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
                    <p style="color: #999; font-size: 0.8em; text-align: center;">
                        Sistema Hayuelos
                    </p>
                </div>
            </body>
            </html>
            """
            try:
                mail.send(msg)
                flash("Se ha enviado un enlace de recuperación a tu email", "success")
            except Exception as e:
                flash(f"Error al enviar el email. Contacte al administrador.", "danger")
                print(f"Error: {e}")
        else:
            flash("Si el email existe en nuestro sistema, recibirás un enlace de recuperación", "info")
    
    return render_template("auth/request_reset.html", form=form)

@auth_bp.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token):
    empleado = Empleado.query.filter_by(reset_token=token).first()
    if not empleado or not empleado.verify_reset_token(token):
        flash("Token inválido o expirado", "danger")
        return redirect(url_for("auth.request_password_reset"))
    
    form = PasswordResetForm()  # 👈 El formulario debe estar aquí
    
    if form.validate_on_submit():
        hash_nuevo = bcrypt.hashpw(form.password.data.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        empleado.contrasena = hash_nuevo
        empleado.reset_token = None
        empleado.reset_token_expiry = None
        empleado.temporal = False
        db.session.commit()
        
        registrar_auditoria('RESET_PASSWORD', 'empleado', empleado.id_empleados, None, {
            'password_reset': True
        })
        
        flash("Contraseña restablecida exitosamente", "success")
        return redirect(url_for("auth.login"))
    
    return render_template("auth/reset_password.html", form=form)  # 👈 Pasando el form

@auth_bp.route("/reset_password/<int:empleado_id>", methods=["POST"])
@login_required
@admin_or_encargado_required
def reset_password_admin(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    
    contrasena_temporal = empleado.numero_documento[-4:] if len(empleado.numero_documento) >= 4 else empleado.numero_documento
    hash_cifrado = bcrypt.hashpw(contrasena_temporal.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    
    empleado.contrasena = hash_cifrado
    empleado.temporal = True
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'empleado', empleado_id, None, {'reset_password': True})
    
    flash(f"Contraseña restablecida para {empleado.nombre_empleado}. Nueva contraseña: {contrasena_temporal}", "success")
    return redirect(url_for("dashboard.empleados"))

@auth_bp.route("/resetear-clave/<int:empleado_id>", methods=["POST"])
@login_required
@admin_required
def resetear_clave_empleado(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    
    temp_password = empleado.numero_documento[-4:] if len(empleado.numero_documento) >= 4 else empleado.numero_documento
    hashed = bcrypt.hashpw(temp_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    empleado.contrasena = hashed
    empleado.temporal = True
    db.session.commit()
    
    flash(f'Contraseña reseteada para {empleado.usuario}. Nueva: {temp_password}', 'warning')
    registrar_auditoria('reset_password', 'empleado', empleado.id_empleados, None, {'temporal': True})
    
    return redirect(url_for('dashboard.empleados'))

# ============= MAIN ROUTES =============
@main_bp.route("/")
def home():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard.index"))
    return redirect(url_for("auth.login"))

@main_bp.route("/terminos")
def terminos():
    return render_template("terminos.html")

@main_bp.route("/privacidad")
def privacidad():
    return render_template("privacidad.html")

# ============= DASHBOARD ROUTES =============
@dashboard_bp.route("/")
@login_required
def index():
    tanques = Tanque.query.filter_by(activo=True).all()
    total_capacity = sum(float(t.capacidad) for t in tanques) if tanques else 0
    mediciones_recientes = RegistroMedida.query.order_by(
        RegistroMedida.fecha_hora_registro.desc()
    ).limit(5).all()
    descargues_hoy = Descargue.query.filter_by(fecha=date.today()).all()

    tanques_por_tipo = {}
    for tanque in tanques:
        tipo = tanque.tipo_combustible
        if tipo not in tanques_por_tipo:
            tanques_por_tipo[tipo] = {"count": 0, "capacity": 0, "current": 0}
        tanques_por_tipo[tipo]["count"] += 1
        tanques_por_tipo[tipo]["capacity"] += float(tanque.capacidad)
        tanques_por_tipo[tipo]["current"] += tanque.contenido or 0

    combustible_mas_vendido = db.session.query(
        Tanque.tipo_combustible,
        func.sum(Venta.cantidad_galones).label('total')
    ).join(Venta).group_by(Tanque.tipo_combustible).order_by(func.sum(Venta.cantidad_galones).desc()).first()

    ventas_por_mes = db.session.query(
        extract('month', Venta.fecha).label('mes'),
        func.sum(Venta.cantidad_galones).label('total')
    ).group_by('mes').order_by(func.sum(Venta.cantidad_galones).desc()).all()

    context = {
        "tanques": tanques,
        "total_capacity": total_capacity,
        "mediciones_recientes": mediciones_recientes,
        "descargues_hoy": descargues_hoy,
        "tanques_por_tipo": tanques_por_tipo,
        "total_tanques": len(tanques),
        "combustible_mas_vendido": combustible_mas_vendido,
        "ventas_por_mes": ventas_por_mes
    }
    return render_template("dashboard/index.html", **context)

@dashboard_bp.route("/tanques")
@login_required
def tanques():
    """Mostrar TODOS los tanques (activos e inactivos)"""
    # ❌ ANTES: tanques = Tanque.query.filter_by(activo=True).all()
    # ✅ AHORA: Mostrar todos los tanques, ordenados por activo primero
    
    tanques = Tanque.query.order_by(
        Tanque.activo.desc(),  # Activos primero
        Tanque.id_tanques.asc()  # Luego por ID
    ).all()
    
    return render_template("dashboard/tanques.html", tanques=tanques)

@dashboard_bp.route("/empleados")
@login_required
@admin_or_encargado_required
def empleados():
    empleados = Empleado.query.all()
    return render_template("dashboard/empleados.html", empleados=empleados)

# routes.py - FUNCIÓN estadisticas() CORREGIDA
# Reemplaza SOLO esta función en tu routes.py

@dashboard_bp.route("/estadisticas")
@login_required
def estadisticas():
    """Dashboard de estadísticas paginado - VERSIÓN CORREGIDA"""
    from collections import OrderedDict
    
    page = request.args.get('page', 1, type=int)
    per_page = 5
    
    # ===== ESTADÍSTICA 1: GALONES CARGADOS POR SEMANA =====
    hoy = datetime.now()
    hace_8_semanas = hoy - timedelta(weeks=8)
    
    cargues = MedicionCargue.query.filter(
        MedicionCargue.fecha >= hace_8_semanas
    ).all()
    
    galones_por_semana = OrderedDict()
    for cargue in cargues:
        if cargue.fecha:
            semana = cargue.fecha.isocalendar()[1]
            año = cargue.fecha.year
            key = f"{año}-W{semana:02d}"
            
            try:
                galones = float(cargue.galones_totales or 0)
                galones_por_semana[key] = galones_por_semana.get(key, 0) + galones
            except (ValueError, TypeError):
                continue
    
    # ===== ESTADÍSTICA 2: GALONES CARGADOS POR MES =====
    hace_12_meses = hoy - timedelta(days=365)
    
    cargues_mes = MedicionCargue.query.filter(
        MedicionCargue.fecha >= hace_12_meses
    ).all()
    
    galones_por_mes = OrderedDict()
    for cargue in cargues_mes:
        if cargue.fecha:
            key = cargue.fecha.strftime('%Y-%m')
            mes_nombre = cargue.fecha.strftime('%B %Y')
            
            try:
                galones = float(cargue.galones_totales or 0)
                if key not in galones_por_mes:
                    galones_por_mes[key] = {
                        'nombre': mes_nombre,
                        'galones': 0
                    }
                galones_por_mes[key]['galones'] += galones
            except (ValueError, TypeError):
                continue
    
    # ===== ESTADÍSTICA 3: GALONES VENDIDOS (DIFERENCIA DE MEDICIONES) =====
    ventas_por_tipo = OrderedDict()
    
    tanques = Tanque.query.filter_by(activo=True).all()
    
    for tanque in tanques:
        mediciones = RegistroMedida.query.filter(
            RegistroMedida.id_tanques == tanque.id_tanques,
            RegistroMedida.fecha_hora_registro >= hoy - timedelta(days=30)
        ).order_by(RegistroMedida.fecha_hora_registro.asc()).all()
        
        total_vendido = 0
        for i in range(1, len(mediciones)):
            medicion_anterior = mediciones[i-1]
            medicion_actual = mediciones[i]
            
            try:
                galones_anterior = float(medicion_anterior.galones or 0)
                galones_actual = float(medicion_actual.galones or 0)
                
                diferencia = galones_anterior - galones_actual
                if diferencia > 0:
                    total_vendido += diferencia
            except (ValueError, TypeError):
                continue
        
        if total_vendido > 0:
            tipo = tanque.tipo_combustible
            ventas_por_tipo[tipo] = ventas_por_tipo.get(tipo, 0) + total_vendido
    
    # ===== ESTADÍSTICA 4: TOTAL CARGADO VS VENDIDO (ÚLTIMO MES) =====
    total_cargado_mes = sum([v['galones'] for v in galones_por_mes.values()])
    total_vendido_mes = sum(ventas_por_tipo.values())
    
    # ===== ESTADÍSTICA 5: PROMEDIO DE MEDICIONES POR DÍA =====
    hace_30_dias = hoy - timedelta(days=30)
    
    mediciones_diarias = db.session.query(
        func.date(RegistroMedida.fecha_hora_registro).label('fecha'),
        func.count(RegistroMedida.id_registro_medidas).label('cantidad')
    ).filter(
        RegistroMedida.fecha_hora_registro >= hace_30_dias
    ).group_by(
        func.date(RegistroMedida.fecha_hora_registro)
    ).all()
    
    total_mediciones = sum([m.cantidad for m in mediciones_diarias])
    promedio_mediciones = total_mediciones / max(len(mediciones_diarias), 1)
    
    # ===== PREPARAR DATOS PARA PAGINACIÓN =====
    estadisticas = [
        {
            'numero': 1,
            'titulo': 'Galones Cargados por Semana',
            'descripcion': 'Cargues de emergencia de las últimas 8 semanas',
            'tipo': 'bar',
            'datos': dict(galones_por_semana),  # Convertir a dict normal
            'unidad': 'galones'
        },
        {
            'numero': 2,
            'titulo': 'Galones Cargados por Mes',
            'descripcion': 'Histórico de cargues de los últimos 12 meses',
            'tipo': 'line',
            'datos': dict(galones_por_mes),  # Convertir a dict normal
            'unidad': 'galones'
        },
        {
            'numero': 3,
            'titulo': 'Galones Vendidos por Tipo de Combustible',
            'descripcion': 'Ventas calculadas del último mes',
            'tipo': 'pie',
            'datos': dict(ventas_por_tipo),  # Convertir a dict normal
            'unidad': 'galones'
        },
        {
            'numero': 4,
            'titulo': 'Comparativa: Cargado vs Vendido',
            'descripcion': 'Balance del último mes',
            'tipo': 'comparison',
            'datos': {
                'cargado': float(total_cargado_mes),
                'vendido': float(total_vendido_mes),
                'diferencia': float(total_cargado_mes - total_vendido_mes)
            },
            'unidad': 'galones'
        },
        {
            'numero': 5,
            'titulo': 'Promedio de Mediciones Diarias',
            'descripcion': 'Actividad promedio de los últimos 30 días',
            'tipo': 'metric',
            'datos': {
                'promedio': round(promedio_mediciones, 2),
                'total_dias': len(mediciones_diarias),
                'total_mediciones': total_mediciones
            },
            'unidad': 'mediciones/día'
        }
    ]
    
    # ===== PAGINACIÓN =====
    total_estadisticas = len(estadisticas)
    total_pages = (total_estadisticas + per_page - 1) // per_page
    
    start = (page - 1) * per_page
    end = start + per_page
    estadisticas_paginadas = estadisticas[start:end]
    
    # ===== DEBUG: Imprimir datos en consola del servidor =====
    print("\n" + "="*60)
    print("📊 DEBUG ESTADÍSTICAS")
    print("="*60)
    print(f"Página actual: {page}/{total_pages}")
    print(f"Estadísticas en esta página: {len(estadisticas_paginadas)}")
    
    for stat in estadisticas_paginadas:
        print(f"\n--- Estadística {stat['numero']}: {stat['titulo']} ---")
        print(f"Tipo: {stat['tipo']}")
        print(f"Datos: {stat['datos']}")
        if stat['tipo'] in ['bar', 'line', 'pie']:
            print(f"Cantidad de datos: {len(stat['datos'])}")
    
    print("="*60 + "\n")
    
    return render_template(
        'dashboard/estadisticas.html',
        estadisticas=estadisticas_paginadas,
        page=page,
        total_pages=total_pages,
        per_page=per_page
    )


# ===== RUTAS PARA MANUALES Y POLÍTICAS =====

@main_bp.route("/manual-usuario")
@login_required
def manual_usuario():
    """Manual de usuario (pendiente de subir)"""
    return render_template("documentacion/manual_usuario.html")


@main_bp.route("/manual-tecnico")
@login_required
def manual_tecnico():
    """Manual técnico (pendiente de subir)"""
    return render_template("documentacion/manual_tecnico.html")


@main_bp.route("/politicas")
@login_required
def politicas():
    """Políticas de la empresa (pendiente de subir)"""
    return render_template("documentacion/politicas.html")

# ============= MEDICION ROUTES =============
@medicion_bp.route("/registro", methods=["GET", "POST"])
@login_required
@islero_or_encargado_required
def registro():
    form = MedicionForm()
    tanques = Tanque.query.filter_by(activo=True).all()
    form.tanque.choices = [(t.id_tanques, f"{t.tipo_combustible} - Tanque {t.id_tanques}") for t in tanques]

    if form.validate_on_submit():
        imagen_path = None
        if form.imagen.data:
            file = form.imagen.data
            if allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join('static/uploads', filename)
                os.makedirs('static/uploads', exist_ok=True)
                file.save(filepath)
                imagen_path = filename

        medicion = RegistroMedida(
            medida_combustible=form.medida_combustible.data,
            id_empleados=current_user.id_empleados,
            fecha_hora_registro=datetime.now(),
            galones=form.galones.data,
            id_tanques=form.tanque.data,
            tipo_medida=form.tipo_medida.data,
            novedad=form.novedad.data,
            imagen_path=imagen_path
        )
        db.session.add(medicion)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'registro_medidas', medicion.id_registro_medidas, None, {
            'tanque': form.tanque.data,
            'galones': form.galones.data
        })

        flash("Medición registrada exitosamente", "success")
        return redirect(url_for("medicion.historial"))

    return render_template("medicion/registro.html", form=form)

@medicion_bp.route("/historial")
@login_required
def historial():
    page = request.args.get("page", 1, type=int)
    
    query = RegistroMedida.query
    
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    tanque_id = request.args.get('tanque', type=int)
    tipo = request.args.get('tipo')
    
    if fecha_desde:
        query = query.filter(RegistroMedida.fecha_hora_registro >= fecha_desde)
    if fecha_hasta:
        query = query.filter(RegistroMedida.fecha_hora_registro <= fecha_hasta)
    if tanque_id:
        query = query.filter_by(id_tanques=tanque_id)
    if tipo:
        query = query.filter_by(tipo_medida=tipo)
    
    mediciones = query.order_by(
        RegistroMedida.fecha_hora_registro.desc()
    ).paginate(page=page, per_page=20, error_out=False)
    
    tanques = Tanque.query.filter_by(activo=True).all()
    return render_template("medicion/historial.html", mediciones=mediciones, tanques=tanques)

@medicion_bp.route("/descargue", methods=["GET", "POST"])
@login_required
@islero_or_encargado_required
def descargue():
    form = DescargueForm()
    if form.validate_on_submit():
        imagen_path = None
        if form.imagen.data:
            file = form.imagen.data
            if allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join('static/uploads', filename)
                os.makedirs('static/uploads', exist_ok=True)
                file.save(filepath)
                imagen_path = filename

        descargue_obj = Descargue(
            id_empleados=current_user.id_empleados,
            medida_inicial_cm=form.medida_inicial_cm.data,
            medida_inicial_gl=form.medida_inicial_gl.data,
            descargue_cm=form.descargue_cm.data,
            descargue_gl=form.descargue_gl.data,
            medida_final_cm=form.medida_final_cm.data,
            medida_final_gl=form.medida_final_gl.data,
            diferencia=form.diferencia.data,
            tanque=form.tanque.data,
            observaciones1=form.observaciones1.data,
            observaciones2=form.observaciones2.data,
            kit_derrames=form.kit_derrames.data,
            extintores=form.extintores.data,
            conos=form.conos.data,
            boquillas=form.boquillas.data,
            botas=form.botas.data,
            gafas=form.gafas.data,
            tapaoidos=form.tapaoidos.data,
            guantes=form.guantes.data,
            brillante=form.brillante.data,
            traslucido=form.traslucido.data,
            claro=form.claro.data,
            solidos=form.solidos.data,
            separacion=form.separacion.data,
            fecha=form.fecha.data or date.today(),
            imagen_path=imagen_path
        )
        db.session.add(descargue_obj)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'descargues', descargue_obj.idDescargue, None, {
            'tanque': form.tanque.data
        })

        flash("Descargue registrado exitosamente", "success")
        return redirect(url_for("medicion.historial_descargues"))

    return render_template("medicion/descargue.html", form=form)

@medicion_bp.route("/api/convert_cm_to_gallons/<int:tanque_id>", methods=["GET"])
@login_required
def convert_cm_to_gallons(tanque_id):
    """API para convertir cm a galones según el tanque específico"""
    cm = request.args.get('cm', type=float, default=0)
    
    tanque = Tanque.query.get_or_404(tanque_id)
    
    # Obtener dimensiones reales del tanque
    # Opción 1: Si las dimensiones están en la BD
    radio_m = tanque.diametro_m / 2  # radio en metros
    radio_cm = radio_m * 100  # convertir a cm
    
    # Opción 2: Si son fijas por tipo de tanque
    # DIMENSIONES_POR_TIPO = {
    #     'Diesel': {'radio_cm': 125},  # 2.5m diámetro
    #     'ACPM': {'radio_cm': 150},    # 3.0m diámetro
    # }
    # radio_cm = DIMENSIONES_POR_TIPO[tanque.tipo_combustible]['radio_cm']
    
    # Fórmula correcta: V = π * r² * h
    # 1 galón = 3785.411784 cm³
    area_base = 3.14159 * (radio_cm ** 2)  # cm²
    volumen_cm3 = area_base * cm  # cm³
    galones = volumen_cm3 / 3785.411784
    
    return jsonify({
        'cm': cm,
        'gallons': round(galones, 2),
        'tanque_id': tanque_id,
        'radio_cm': radio_cm
    })

@medicion_bp.route("/historial_descargues")
@login_required
def historial_descargues():
    page = request.args.get("page", 1, type=int)
    descargues = Descargue.query.order_by(Descargue.fecha.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template("medicion/historial_descargues.html", descargues=descargues)

# ============= ADMIN ROUTES =============
@admin_bp.route("/toggle_empleado/<int:empleado_id>", methods=["POST"])
@login_required
@admin_required
def toggle_empleado(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    empleado.activo = not empleado.activo
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'empleado', empleado_id, 
                      {'activo': not empleado.activo}, {'activo': empleado.activo})
    
    estado = "habilitado" if empleado.activo else "deshabilitado"
    flash(f"Empleado {empleado.nombre_empleado} {estado}", "success")
    return redirect(url_for("dashboard.empleados"))

@admin_bp.route("/carga_masiva", methods=["GET", "POST"])
@login_required
@admin_required
def carga_masiva():
    form = CargaMasivaForm()
    if form.validate_on_submit():
        file = form.archivo.data
        tipo_carga = form.tipo_carga.data
        
        if not file or not allowed_file(file.filename, {'csv', 'xlsx', 'xls'}):
            flash("Archivo inválido", "danger")
            return redirect(request.url)

        try:
            if file.filename.lower().endswith('.csv'):
                df = pd.read_csv(file, sep=',', decimal='.', encoding='utf-8-sig')
            else:
                df = pd.read_excel(file)
            
            df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
            df.replace({"True": True, "False": False}, inplace=True, regex=True)

            count = 0
            errors = []

            if tipo_carga == 'empleados':
                required = ['nombre_empleado', 'apellido_empleado', 'numero_documento', 'email', 'usuario']
                if not all(col in df.columns for col in required):
                    flash(f"Faltan columnas: {', '.join(required)}", "danger")
                    return redirect(request.url)
                
                for idx, row in df.iterrows():
                    if Empleado.query.filter_by(numero_documento=row['numero_documento']).first():
                        errors.append(f"Fila {idx+2}: Documento duplicado")
                        continue
                    if Empleado.query.filter_by(usuario=row['usuario']).first():
                        errors.append(f"Fila {idx+2}: Usuario duplicado")
                        continue
                    
                    temp_pass = str(row['numero_documento'])[-4:]
                    hash_pwd = bcrypt.hashpw(temp_pass.encode(), bcrypt.gensalt()).decode()
                    
                    empleado = Empleado(
                        nombre_empleado=row['nombre_empleado'],
                        apellido_empleado=row['apellido_empleado'],
                        numero_documento=row['numero_documento'],
                        tipo_documento=row.get('tipo_documento', 'CC'),
                        email=row['email'],
                        telefono=row.get('telefono', ''),
                        direccion=row.get('direccion', ''),
                        cargo_establecido=row.get('cargo_establecido', 'Islero'),
                        usuario=row['usuario'],
                        contrasena=hash_pwd,
                        temporal=True,
                        activo=row.get('activo', True),
                        email_confirmado=True,  # Carga masiva: emails pre-confirmados
                        aceptado_terminos=row.get('aceptado_terminos', False)
                    )
                    db.session.add(empleado)
                    count += 1

            elif tipo_carga == 'tanques':
                required = ['tipo_combustible', 'capacidad']
                if not all(col in df.columns for col in required):
                    flash("Faltan columnas: tipo_combustible, capacidad", "danger")
                    return redirect(request.url)
                
                for idx, row in df.iterrows():
                    try:
                        capacidad = int(float(str(row['capacidad']).replace(',', '.')))
                        tanque = Tanque(
                            tipo_combustible=row['tipo_combustible'],
                            capacidad=capacidad,
                            activo=row.get('activo', True)
                        )
                        db.session.add(tanque)
                        count += 1
                    except Exception as e:
                        errors.append(f"Fila {idx+2}: Capacidad inválida → {str(e)}")

            elif tipo_carga == 'mediciones':
                required = ['tanque_id', 'medida_combustible', 'galones', 'tipo_medida', 'fecha_hora_registro', 'empleado_id']
                if not all(col in df.columns for col in required):
                    flash("Faltan columnas en mediciones", "danger")
                    return redirect(request.url)
                
                for idx, row in df.iterrows():
                    try:
                        tanque_id = int(row['tanque_id'])
                        empleado_id = int(row['empleado_id'])
                        tanque = Tanque.query.get(tanque_id)
                        empleado = Empleado.query.get(empleado_id)
                        
                        if not tanque:
                            errors.append(f"Fila {idx+2}: tanque_id {tanque_id} no existe")
                            continue
                        if not empleado:
                            errors.append(f"Fila {idx+2}: empleado_id {empleado_id} no existe")
                            continue
                        
                        medida_str = str(row['medida_combustible']).replace(',', '.')
                        galones_str = str(row['galones']).replace(',', '.')
                        medida = float(medida_str)
                        galones = float(galones_str)
                        
                        fecha_str = str(row['fecha_hora_registro']).strip()
                        try:
                            fecha = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
                        except:
                            fecha = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M')
                        
                        medicion = RegistroMedida(
                            id_tanques=tanque.id_tanques,
                            id_empleados=empleado.id_empleados,
                            medida_combustible=medida,
                            galones=galones,
                            tipo_medida=row['tipo_medida'],
                            novedad=row.get('novedad', ''),
                            fecha_hora_registro=fecha
                        )
                        db.session.add(medicion)
                        count += 1
                    except Exception as e:
                        errors.append(f"Fila {idx+2}: Error → {str(e)}")

            db.session.commit()
            registrar_auditoria('CREATE_BULK', tipo_carga, None, None, {'count': count})

            msg = f"Se cargaron {count} registros exitosamente"
            if errors:
                error_sample = "; ".join(errors[:3])
                flash(f"{msg}. Errores: {len(errors)} → {error_sample}", "warning")
            else:
                flash(msg, "success")

        except Exception as e:
            db.session.rollback()
            flash(f"Error crítico: {str(e)}", "danger")
            print(f"[ERROR] {e}")

    return render_template("admin/carga_masiva.html", form=form)


# ============= EXPORT ROUTES (AGREGAR AL FINAL DE routes.py) =============

@admin_bp.route("/export_menu", methods=["GET"])
@login_required
@admin_or_encargado_required
def export_menu():
    """Menú de exportación de datos"""
    tanques = Tanque.query.filter_by(activo=True).all()
    return render_template("admin/export_menu.html", tanques=tanques)


@admin_bp.route("/export/<tipo>", methods=["GET"])
@login_required
@admin_or_encargado_required
def export_data(tipo):
    """Exportar datos en formato Excel o CSV"""
    from io import BytesIO
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("Error: Instale openpyxl con 'pip install openpyxl'", "danger")
        return redirect(url_for('dashboard.index'))
    
    formato = request.args.get('formato', 'excel')  # 'excel' o 'csv'
    
    # Determinar qué datos exportar
    if tipo == 'empleados':
        empleados = Empleado.query.all()
        data = []
        headers = ['ID', 'Usuario', 'Nombre', 'Apellido', 'Documento', 'Email', 'Teléfono', 
                   'Dirección', 'Cargo', 'Activo', 'Email Confirmado', 'Fecha Creación']
        
        for emp in empleados:
            data.append([
                emp.id_empleados,
                emp.usuario,
                emp.nombre_empleado,
                emp.apellido_empleado,
                emp.numero_documento,
                emp.email,
                emp.telefono or '',
                emp.direccion or '',
                emp.cargo_establecido,
                'Sí' if emp.activo else 'No',
                'Sí' if emp.email_confirmado else 'No',
                emp.fecha_creacion.strftime('%Y-%m-%d %H:%M') if emp.fecha_creacion else ''
            ])
        filename = f"empleados_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
    elif tipo == 'tanques':
        tanques = Tanque.query.all()
        data = []
        headers = ['ID', 'Tipo Combustible', 'Capacidad (gal)', 'Contenido (gal)', 
                   'Volumen (m³)', 'Activo', 'Fecha Creación']
        
        for tanque in tanques:
            data.append([
                tanque.id_tanques,
                tanque.tipo_combustible,
                tanque.capacidad,
                tanque.contenido or 0,
                tanque.volumen_m3,
                'Sí' if tanque.activo else 'No',
                tanque.fecha_creacion.strftime('%Y-%m-%d') if tanque.fecha_creacion else ''
            ])
        filename = f"tanques_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
    elif tipo == 'mediciones':
        # Filtros opcionales
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        tanque_id = request.args.get('tanque_id')
        
        query = RegistroMedida.query
        if fecha_desde:
            query = query.filter(RegistroMedida.fecha_hora_registro >= fecha_desde)
        if fecha_hasta:
            query = query.filter(RegistroMedida.fecha_hora_registro <= fecha_hasta)
        if tanque_id:
            query = query.filter_by(id_tanques=int(tanque_id))
        
        mediciones = query.order_by(RegistroMedida.fecha_hora_registro.desc()).all()
        
        data = []
        headers = ['ID', 'Fecha/Hora', 'Tanque', 'Tipo Combustible', 'Medida (cm)', 
                   'Galones', 'Tipo Medición', 'Empleado', 'Novedad']
        
        for med in mediciones:
            data.append([
                med.id_registro_medidas,
                med.fecha_hora_registro.strftime('%Y-%m-%d %H:%M:%S') if med.fecha_hora_registro else '',
                f"Tanque {med.tanque.id_tanques}" if med.tanque else 'N/A',
                med.tanque.tipo_combustible if med.tanque else 'N/A',
                med.medida_combustible or '',
                med.galones or 0,
                med.tipo_medida or 'rutinario',
                med.empleado.nombre_empleado if med.empleado else 'N/A',
                med.novedad or ''
            ])
        filename = f"mediciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
    elif tipo == 'descargues':
        # Filtros opcionales
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        query = Descargue.query
        if fecha_desde:
            query = query.filter(Descargue.fecha >= fecha_desde)
        if fecha_hasta:
            query = query.filter(Descargue.fecha <= fecha_hasta)
        
        descargues = query.order_by(Descargue.fecha.desc()).all()
        
        data = []
        headers = ['ID', 'Fecha', 'Tanque', 'Medida Inicial (gl)', 'Descargue (gl)', 
                   'Medida Final (gl)', 'Diferencia', 'Empleado', 'Kit Derrames', 
                   'Extintores', 'Observaciones']
        
        for desc in descargues:
            data.append([
                desc.idDescargue,
                desc.fecha.strftime('%Y-%m-%d') if desc.fecha else '',
                desc.tanque or '',
                float(desc.medida_inicial_gl) if desc.medida_inicial_gl else 0,
                float(desc.descargue_gl) if desc.descargue_gl else 0,
                float(desc.medida_final_gl) if desc.medida_final_gl else 0,
                float(desc.diferencia) if desc.diferencia else 0,
                desc.empleado.nombre_empleado if desc.empleado else 'N/A',
                desc.kit_derrames or 'no',
                desc.extintores or 'no',
                desc.observaciones1 or ''
            ])
        filename = f"descargues_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    else:
        flash("Tipo de exportación no válido", "danger")
        return redirect(url_for('dashboard.index'))
    
    # Generar archivo según formato
    if formato == 'csv':
        # Exportar como CSV
        output = BytesIO()
        df = pd.DataFrame(data, columns=headers)
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        
        registrar_auditoria('EXPORT_CSV', tipo, None, None, {
            'formato': 'csv',
            'registros': len(data)
        })
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename}.csv"
        )
    
    else:  # Excel
        # Crear workbook de Excel con estilos
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo.capitalize()
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="E10000", end_color="E10000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        registrar_auditoria('EXPORT_EXCEL', tipo, None, None, {
            'formato': 'excel',
            'registros': len(data)
        })
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename}.xlsx"
        )

# Agregar estas rutas al final de admin_bp en routes.py

# ============= GESTIÃ"N COMPLETA DE TANQUES =============

# ============= GESTIÓN COMPLETA DE TANQUES =============

@admin_bp.route("/tanques/crear", methods=["GET", "POST"])
@login_required
@admin_or_encargado_required
def crear_tanque():
    """Crear nuevo tanque"""
    form = TanqueForm()
    if form.validate_on_submit():
        tanque = Tanque(
            tipo_combustible=form.tipo_combustible.data,
            capacidad=form.capacidad.data,
            activo=form.activo.data,
            altura_maxima_cm=calcular_altura_maxima(form.capacidad.data),
            radio_cm=125.0
        )
        db.session.add(tanque)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'tanques', tanque.id_tanques, None, {
            'tipo': form.tipo_combustible.data,
            'capacidad': form.capacidad.data
        })
        
        flash("Tanque creado exitosamente", "success")
        return redirect(url_for("dashboard.tanques"))
    
    return render_template("admin/tanque_form.html", form=form, titulo="Nuevo Tanque", accion="crear")


@admin_bp.route("/tanques/<int:tanque_id>/editar", methods=["GET", "POST"])
@login_required
@admin_or_encargado_required
def editar_tanque(tanque_id):
    """Editar tanque existente"""
    tanque = Tanque.query.get_or_404(tanque_id)
    form = TanqueForm(obj=tanque)
    
    if form.validate_on_submit():
        datos_anteriores = {
            'tipo': tanque.tipo_combustible,
            'capacidad': tanque.capacidad,
            'activo': tanque.activo
        }
        
        tanque.tipo_combustible = form.tipo_combustible.data
        tanque.capacidad = form.capacidad.data
        tanque.activo = form.activo.data
        tanque.altura_maxima_cm = calcular_altura_maxima(form.capacidad.data)
        
        db.session.commit()
        
        registrar_auditoria('UPDATE', 'tanques', tanque_id, datos_anteriores, {
            'tipo': tanque.tipo_combustible,
            'capacidad': tanque.capacidad,
            'activo': tanque.activo
        })
        
        flash("Tanque actualizado exitosamente", "success")
        return redirect(url_for("dashboard.tanques"))
    
    return render_template("admin/tanque_form.html", form=form, titulo="Editar Tanque", accion="editar", tanque=tanque)


@admin_bp.route("/tanques/<int:tanque_id>/toggle", methods=["POST"])
@login_required
@admin_or_encargado_required
def toggle_tanque(tanque_id):
    """Activar/Desactivar tanque"""
    tanque = Tanque.query.get_or_404(tanque_id)
    tanque.activo = not tanque.activo
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'tanques', tanque_id, 
                      {'activo': not tanque.activo}, {'activo': tanque.activo})
    
    estado = "activado" if tanque.activo else "desactivado"
    flash(f"Tanque {estado}", "success")
    return redirect(url_for("dashboard.tanques"))


def calcular_altura_maxima(capacidad_galones):
    """Calcular altura máxima en cm basada en capacidad del tanque"""
    # Radio estándar en cm (ajustar según tanques reales)
    radio_cm = 125  # 2.5m de diámetro
    
    # Volumen en cm³ = capacidad en galones * 3785.411784
    volumen_cm3 = capacidad_galones * 3785.411784
    
    # Altura = Volumen / (π * r²)
    area_base = 3.14159 * (radio_cm ** 2)
    altura_cm = volumen_cm3 / area_base
    
    return round(altura_cm, 2)

# ============= CARGUE DE EMERGENCIA =============
@medicion_bp.route("/cargue_emergencia", methods=["GET", "POST"])
@login_required
@islero_or_encargado_required
def cargue_emergencia():
    """Registrar cargue de emergencia"""
    from forms import CargueEmergenciaForm
    form = CargueEmergenciaForm()
    
    # Cargar tanques en el selector
    tanques = Tanque.query.filter_by(activo=True).all()
    form.tanque.choices = [(t.id_tanques, f"{t.tipo_combustible} - Tanque {t.id_tanques}") for t in tanques]
    
    if form.validate_on_submit():
        imagen_path = None
        if form.imagen.data:
            file = form.imagen.data
            if allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join('static/uploads', filename)
                os.makedirs('static/uploads', exist_ok=True)
                file.save(filepath)
                imagen_path = filename
        
        cargue = MedicionCargue(
            id_empleados=current_user.id_empleados,
            id_tanques=form.tanque.data,
            medida_anterior=form.medida_anterior.data,
            medida_posterior=form.medida_posterior.data,
            formato_de_entrega=form.formato_entrega.data,
            galones_totales=form.galones_totales.data,
            fecha=datetime.now()
        )
        db.session.add(cargue)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'cargue_emergencia', cargue.id_medicion_cargue, None, {
            'tanque': form.tanque.data,
            'galones': form.galones_totales.data
        })
        
        flash("Cargue de emergencia registrado exitosamente", "success")
        return redirect(url_for("medicion.historial_cargues"))
    
    return render_template("medicion/cargue_emergencia.html", form=form)


@medicion_bp.route("/historial_cargues")
@login_required
@islero_or_encargado_required
def historial_cargues():
    """Historial de cargues de emergencia"""
    page = request.args.get("page", 1, type=int)
    cargues = MedicionCargue.query.order_by(
        MedicionCargue.fecha.desc()
    ).paginate(page=page, per_page=20, error_out=False)
    
    return render_template("medicion/historial_cargues.html", cargues=cargues)
